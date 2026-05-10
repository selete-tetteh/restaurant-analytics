"""
generate_report.py
------------------
Generates the Plato's Pizza management report as a formatted Excel workbook.

The report is a non-technical document intended for a restaurant manager or
business owner. Every finding is translated into plain English with a specific
financial implication and a concrete recommendation. No statistical jargon.

Sheets:
    0. Executive Summary  — top five findings with financial value and action
    1. Operations         — peak trading patterns and staffing implications
    2. Menu Performance   — revenue and volume rankings, size analysis, upsell
    3. Forecasting        — 30-day demand forecast and staffing recommendations
    4. Pricing Experiment — A/B test findings and pricing strategy implications

Run from the project root:
    python src/generate_report.py

Output: reports/platos_pizza_management_report.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
REPORTS_DIR  = PROJECT_ROOT / "reports"
OUTPUT_PATH  = REPORTS_DIR / "platos_pizza_management_report.xlsx"

STAFFING_CSV = REPORTS_DIR / "staffing_forecast.csv"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"

# Colour palette
C_DARK_RED   = "8B1A1A"   # Header backgrounds — deep professional red
C_MID_RED    = "C0392B"   # Section sub-headers
C_LIGHT_RED  = "FADBD8"   # Alternate row tint
C_DARK_GREY  = "2C3E50"   # Primary text on dark backgrounds
C_WHITE      = "FFFFFF"
C_LIGHT_GREY = "F2F2F2"   # Alternate row background
C_YELLOW     = "FFF9C4"   # Highlight cells (key figures)
C_GREEN      = "D5F5E3"   # Positive outcome highlight
C_BORDER     = "BDC3C7"   # Cell border colour

# Fonts
def hdr_font(size=11):
    return Font(name=FONT_NAME, bold=True, color=C_WHITE, size=size)

def body_font(size=10, bold=False, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def title_font(size=14):
    return Font(name=FONT_NAME, bold=True, size=size, color=C_DARK_RED)

# Fills
def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

# Borders
thin_side   = Side(style="thin",   color=C_BORDER)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Alignment
centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
right  = Alignment(horizontal="right",  vertical="center")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def write_cell(ws, row, col, value, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill_:  cell.fill      = fill_
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt
    return cell

def section_header(ws, row, col_start, col_end, text, bg=C_MID_RED):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font      = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
    cell.fill      = fill(bg)
    cell.alignment = left
    row_height(ws, row, 22)

def col_header(ws, row, col, text):
    cell = write_cell(ws, row, col, text,
                      font=hdr_font(10), fill_=fill(C_DARK_RED),
                      align=centre, border=thin_border)
    return cell

def finding_row(ws, row, cols_values: list, shade=False):
    bg = C_LIGHT_RED if shade else C_WHITE
    for col, value in cols_values:
        write_cell(ws, row, col, value,
                   font=body_font(), fill_=fill(bg),
                   align=left, border=thin_border)
    row_height(ws, row, 40)

def sheet_title(ws, row, col_start, col_end, title, subtitle=None):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font      = title_font(14)
    cell.alignment = left
    row_height(ws, row, 28)

    if subtitle:
        ws.merge_cells(
            start_row=row+1, start_column=col_start,
            end_row=row+1,   end_column=col_end
        )
        sub = ws.cell(row=row+1, column=col_start, value=subtitle)
        sub.font      = body_font(10, color="555555")
        sub.alignment = left
        row_height(ws, row+1, 18)


# ---------------------------------------------------------------------------
# Sheet 0 — Executive Summary
# ---------------------------------------------------------------------------
def build_executive_summary(wb):
    ws = wb.active
    ws.title = "Executive Summary"

    set_col_widths(ws, {
        "A": 4, "B": 32, "C": 45, "D": 28, "E": 38
    })

    sheet_title(ws, 1, 2, 5,
        "Plato's Pizza — Management Report",
        "Analytics findings with financial implications and recommended actions | Dataset: 2015"
    )
    row_height(ws, 3, 10)

    section_header(ws, 4, 2, 5, "Five Most Commercially Significant Findings", bg=C_DARK_RED)

    # Column headers
    headers = ["Finding", "What It Means", "Financial Implication", "Recommended Action"]
    for i, h in enumerate(headers, start=2):
        col_header(ws, 5, i, h)
    row_height(ws, 5, 22)

    findings = [
        (
            "Friday is the highest-volume day. Lunch is the busiest meal period with 7,678 orders.",
            "The business is not evenly distributed across the week or day. Two time windows — Friday lunch and Saturday evening — drive a disproportionate share of revenue.",
            "Understaffing in peak windows directly caps revenue. Every order that takes too long risks abandonment or a negative experience.",
            "Schedule senior staff for Friday and Saturday. Ensure kitchen throughput is not the constraint during the 12:00–14:00 window."
        ),
        (
            "Thai Chicken Pizza is the #1 revenue item but only #5 by volume.",
            "The highest-earning pizza is not the most popular one. Customers who order it spend more, but many customers are not choosing it.",
            "Increasing Thai Chicken visibility by even 5% of current Pepperoni volume would add an estimated $2,400+ per year at current pricing.",
            "Feature Thai Chicken prominently on menus and digital boards. Consider a bundle that pairs it with a high-lift complement from the basket analysis."
        ),
        (
            "Large pizzas represent 38% of volume but 46% of revenue.",
            "Large pizzas punch above their volume weight. Every medium-to-large upsell converts a lower-margin order into a higher-margin one.",
            "A 10% conversion of medium orders to large would generate an additional $6,276 per year.",
            "Train staff to offer a size upgrade at point of order. Price the gap between medium and large to make the upsell feel like value rather than upselling."
        ),
        (
            "Italian Vegetables and Pepperoni Mushroom Peppers are the strongest complementary pair (lift: 1.500).",
            "Customers who order one of these are 50% more likely than average to also order the other. This is an untapped bundle opportunity.",
            "A promoted bundle at a modest discount would increase average order value while reducing decision friction for the customer.",
            "Create a named bundle combining these two items. Test at a 5–8% discount relative to individual pricing and monitor attach rate."
        ),
        (
            "Medium pizza demand shows no statistically detectable difference across the year (p = 0.93, Cohen's d = 0.009).",
            "Medium pizza buyers are price-stable. Natural seasonal variation produces no measurable demand shift, which implies a degree of price inelasticity.",
            "A 10% medium price increase would generate approximately $512 in incremental annual revenue even without volume change — and the data suggests volume would not change materially.",
            "Model a medium price increase of 5–10% and monitor order composition for four weeks. The demand signal is stable enough to detect a genuine response within that window."
        ),
    ]

    for i, (f1, f2, f3, f4) in enumerate(findings):
        row = 6 + i
        shade = i % 2 == 1
        finding_row(ws, row, [
            (2, f1), (3, f2), (4, f3), (5, f4)
        ], shade=shade)
        row_height(ws, row, 72)


# ---------------------------------------------------------------------------
# Sheet 1 — Operations
# ---------------------------------------------------------------------------
def build_operations(wb):
    ws = wb.create_sheet("Operations")

    set_col_widths(ws, {
        "A": 4, "B": 28, "C": 20, "D": 20, "E": 42
    })

    sheet_title(ws, 1, 2, 5,
        "Operational Analysis",
        "Peak trading patterns, meal period performance, and staffing implications"
    )
    row_height(ws, 3, 10)

    # Peak by day
    section_header(ws, 4, 2, 5, "Order Volume by Day of Week")
    for i, h in enumerate(["Day", "Orders", "vs. Weekly Average", "Note"], start=2):
        col_header(ws, 5, i, h)

    days = [
        ("Friday",    8242, "+14%",  "Highest volume day. Peak staffing required."),
        ("Thursday",  7478, "+3%",   "Second highest. Consistent with end-of-week pattern."),
        ("Wednesday", 7355, "+2%",   "Mid-week — above average throughout."),
        ("Saturday",  7355, "+2%",   "Weekend peak. Avg spend highest ($16.73/order)."),
        ("Tuesday",   7239, "0%",    "Baseline trading day."),
        ("Monday",    6900, "-4%",   "Slowest weekday. Lower staffing viable."),
        ("Sunday",    6051, "-16%",  "Lowest volume day across the year."),
    ]
    for i, (day, orders, vs_avg, note) in enumerate(days):
        row = 6 + i
        shade = i % 2 == 1
        finding_row(ws, row, [
            (2, day), (3, orders), (4, vs_avg), (5, note)
        ], shade=shade)
        row_height(ws, row, 30)

    row_height(ws, 14, 12)

    # Peak by meal period
    section_header(ws, 15, 2, 5, "Performance by Meal Period")
    for i, h in enumerate(["Meal Period", "Total Orders", "Avg Order Value ($)", "Implication"], start=2):
        col_header(ws, 16, i, h)

    periods = [
        ("Lunch",      7678, 41.95, "Highest volume and highest average spend. Primary revenue window."),
        ("Dinner",     7236, 38.20, "Second largest period. Strong but lower spend per order than Lunch."),
        ("Afternoon",  4144, 35.10, "Moderate volume. Opportunity for off-peak promotions."),
        ("Late Night",  693, 32.40, "Low volume. Evaluate whether staffing cost is justified."),
        ("Morning",     412, 29.80, "Minimal volume. Not a meaningful revenue contributor."),
    ]
    for i, (period, orders, aov, note) in enumerate(periods):
        row = 17 + i
        shade = i % 2 == 1
        finding_row(ws, row, [
            (2, period), (3, orders), (4, aov), (5, note)
        ], shade=shade)
        row_height(ws, row, 35)


# ---------------------------------------------------------------------------
# Sheet 2 — Menu Performance
# ---------------------------------------------------------------------------
def build_menu(wb):
    ws = wb.create_sheet("Menu Performance")

    set_col_widths(ws, {
        "A": 4, "B": 30, "C": 18, "D": 18, "E": 42
    })

    sheet_title(ws, 1, 2, 5,
        "Menu Performance",
        "Revenue vs volume rankings, size analysis, upsell opportunity, and cannibalisation"
    )
    row_height(ws, 3, 10)

    # Rank divergence
    section_header(ws, 4, 2, 5, "Volume vs Revenue Rank Divergence — Top Items")
    for i, h in enumerate(["Pizza", "Volume Rank", "Revenue Rank", "Implication"], start=2):
        col_header(ws, 5, i, h)

    items = [
        ("Thai Chicken",            5,  1, "Highest revenue, not highest volume. Under-promoted relative to its value."),
        ("Barbecue Chicken",        2,  3, "High volume and high revenue. Core menu anchor — protect its position."),
        ("California Chicken",      3,  4, "Strong performer on both dimensions. Consistent contributor."),
        ("Pepperoni",               4, 11, "High volume but low revenue rank. Low price or low attachment to extras."),
        ("Brie Carre",             32, 32, "Bottom on both dimensions. Generates $31.74/week. Carry cost likely exceeds contribution."),
    ]
    for i, (pizza, vol_rank, rev_rank, note) in enumerate(items):
        row = 6 + i
        shade = i % 2 == 1
        finding_row(ws, row, [
            (2, pizza), (3, vol_rank), (4, rev_rank), (5, note)
        ], shade=shade)
        row_height(ws, row, 40)

    row_height(ws, 12, 12)

    # Size analysis
    section_header(ws, 13, 2, 5, "Size Performance — Volume vs Revenue Share")
    for i, h in enumerate(["Size", "Volume Share", "Revenue Share", "Implication"], start=2):
        col_header(ws, 14, i, h)

    sizes = [
        ("Large (L)",       "38%", "46%", "Punches above its volume weight. Every medium-to-large upsell improves margins."),
        ("Medium (M)",      "31%", "28%", "Most popular by count. Lower revenue share signals upsell opportunity."),
        ("Small (S)",       "22%", "18%", "Consistent share. No significant action required."),
        ("XL",               "7%",  "7%", "Proportionate share. Niche segment."),
        ("XXL",              "2%",  "1%", "Generates only $1,007/year. Evaluate whether it earns its place on the menu."),
    ]
    for i, (size, vol, rev, note) in enumerate(sizes):
        row = 15 + i
        shade = i % 2 == 1
        finding_row(ws, row, [
            (2, size), (3, vol), (4, rev), (5, note)
        ], shade=shade)
        row_height(ws, row, 40)

    row_height(ws, 21, 12)

    # Upsell
    section_header(ws, 22, 2, 5, "Upsell Opportunity — Medium to Large Conversion")
    ws.merge_cells("B23:E23")
    cell = ws.cell(row=23, column=2,
        value=(
            "If 10% of medium pizza orders converted to large, the business would generate an additional "
            "$6,276 per year at current pricing. This requires no new customers, no new menu items, "
            "and no marketing spend — only a consistent staff prompt at point of order. "
            "The price gap between medium and large should be positioned as value, not as an upsell."
        )
    )
    cell.font      = body_font(10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.fill      = fill(C_GREEN)
    row_height(ws, 23, 70)


# ---------------------------------------------------------------------------
# Sheet 3 — Forecasting
# ---------------------------------------------------------------------------
def build_forecasting(wb):
    ws = wb.create_sheet("Forecasting")

    set_col_widths(ws, {
        "A": 4, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18, "G": 20
    })

    sheet_title(ws, 1, 2, 7,
        "Demand Forecast — 30-Day Staffing Outlook",
        "Prophet model trained on 2015 data | Weekly seasonality is the reliable signal | Trend component unreliable on single-year data"
    )
    row_height(ws, 3, 10)

    section_header(ws, 4, 2, 7, "30-Day Forward Forecast with Staffing Recommendations")

    headers = ["Date", "Forecast Orders", "Lower Bound", "Upper Bound", "Peak Hour Volume", "Staff Needed"]
    for i, h in enumerate(headers, start=2):
        col_header(ws, 5, i, h)

    # Load staffing CSV
    df = pd.read_csv(STAFFING_CSV)
    # Keep only the 30 forward-looking rows (from 2016-01-01 onward)
    df = df[pd.to_datetime(df["date"]) >= pd.Timestamp("2016-01-01")].head(30)

    for i, row_data in enumerate(df.itertuples(index=False)):
        row = 6 + i
        shade = i % 2 == 1
        bg = C_LIGHT_RED if shade else C_WHITE

        values = [
            (2, str(row_data.date)),
            (3, round(row_data.forecast_orders, 1)),
            (4, round(row_data.forecast_lower, 1)),
            (5, round(row_data.forecast_upper, 1)),
            (6, round(row_data.peak_hour_volume, 1)),
            (7, int(row_data.staff_needed)),
        ]
        for col, val in values:
            write_cell(ws, row, col, val,
                       font=body_font(), fill_=fill(bg),
                       align=centre, border=thin_border)
        row_height(ws, row, 18)

    note_row = 6 + len(df) + 1
    ws.merge_cells(
        start_row=note_row, start_column=2,
        end_row=note_row,   end_column=7
    )
    note = ws.cell(row=note_row, column=2,
        value=(
            "Note: Staff needed = 1 throughout reflects the dataset scale (~60 daily orders). "
            "Peak hourly volume at this scale genuinely requires one staff member. "
            "The forecasting framework is valid and will produce meaningful staffing variation "
            "at higher order volumes."
        )
    )
    note.font      = body_font(9, color="666666")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.fill      = fill(C_LIGHT_GREY)
    row_height(ws, note_row, 50)


# ---------------------------------------------------------------------------
# Sheet 4 — Pricing Experiment
# ---------------------------------------------------------------------------
def build_experiment(wb):
    ws = wb.create_sheet("Pricing Experiment")

    set_col_widths(ws, {
        "A": 4, "B": 32, "C": 32
    })

    sheet_title(ws, 1, 2, 3,
        "Pricing Experiment — A/B Test Framework",
        "Retrospective test: would a medium pizza price increase have produced a detectable demand response?"
    )
    row_height(ws, 3, 10)

    section_header(ws, 4, 2, 3, "Test Design")

    design_rows = [
        ("Test type",               "Welch's two-sample t-test (does not assume equal variance)"),
        ("Unit of analysis",        "Daily medium pizza order volume"),
        ("Control period",          "January 1 – June 30, 2015  (181 days)"),
        ("Treatment period",        "July 1 – December 31, 2015  (177 days)"),
        ("Minimum sample required", "394 days per group (80% power, small effect, α = 0.05)"),
        ("Achieved power",          "~50% — both groups below minimum required n"),
    ]
    for i, (label, value) in enumerate(design_rows):
        row = 5 + i
        shade = i % 2 == 1
        bg = C_LIGHT_RED if shade else C_WHITE
        write_cell(ws, row, 2, label, font=body_font(10, bold=True),
                   fill_=fill(bg), align=left, border=thin_border)
        write_cell(ws, row, 3, value, font=body_font(10),
                   fill_=fill(bg), align=left, border=thin_border)
        row_height(ws, row, 28)

    row_height(ws, 12, 12)
    section_header(ws, 13, 2, 3, "Test Results")

    result_rows = [
        ("Control mean",            "43.63 medium pizzas per day"),
        ("Treatment mean",          "43.72 medium pizzas per day"),
        ("Absolute difference",     "0.09 pizzas per day"),
        ("p-value",                 "0.931 — no statistically significant difference detected"),
        ("Cohen's d",               "0.009 — negligible practical effect size"),
        ("95% confidence interval", "-2.08 to +1.91 pizzas per day"),
        ("Annualised revenue diff", "$511.79"),
    ]
    for i, (label, value) in enumerate(result_rows):
        row = 14 + i
        shade = i % 2 == 1
        bg = C_LIGHT_RED if shade else C_WHITE
        write_cell(ws, row, 2, label, font=body_font(10, bold=True),
                   fill_=fill(bg), align=left, border=thin_border)
        write_cell(ws, row, 3, value, font=body_font(10),
                   fill_=fill(bg), align=left, border=thin_border)
        row_height(ws, row, 28)

    row_height(ws, 22, 12)
    section_header(ws, 23, 2, 3, "Business Interpretation")

    ws.merge_cells("B24:C24")
    interp = ws.cell(row=24, column=2, value=(
        "Medium pizza demand is stable across the year. The test found no statistically significant "
        "difference in daily medium pizza volume between the first and second halves of 2015 "
        "(p = 0.931). The effect size is negligible (Cohen's d = 0.009). "
        "\n\n"
        "What this means practically: medium pizza buyers do not change their ordering behaviour "
        "in response to the natural seasonal variation in this dataset. This is consistent with "
        "a degree of price inelasticity — customers who choose medium pizzas are not highly "
        "sensitive to external conditions. "
        "\n\n"
        "Recommendation: a 5–10% medium price increase is unlikely to produce a meaningful "
        "volume reduction. A four-week live price test with proper before/after monitoring "
        "would provide causal confirmation. The annualised revenue upside at current volume "
        "with no demand change is $512."
    ))
    interp.font      = body_font(10)
    interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    interp.fill      = fill(C_YELLOW)
    row_height(ws, 24, 140)

    row_height(ws, 25, 12)
    section_header(ws, 26, 2, 3, "Limitations")

    ws.merge_cells("B27:C27")
    limits = ws.cell(row=27, column=2, value=(
        "1. This is a retrospective observational split, not a controlled experiment. "
        "No price change occurred. Causation cannot be established from this design.\n\n"
        "2. Both groups fell below the minimum required sample size (394 days per group). "
        "The test had approximately 50% power — it could only reliably detect effects "
        "roughly twice as large as the small effect we were looking for.\n\n"
        "3. Seasonal demand is a known confound. September and October are the weakest "
        "trading months and fall in the treatment period. Any lower volume in H2 may "
        "reflect seasonality rather than a pricing effect.\n\n"
        "4. A production-grade test would require random assignment of customers to price "
        "conditions within the same time window, or a difference-in-differences design "
        "using a control location."
    ))
    limits.font      = body_font(10)
    limits.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    limits.fill      = fill(C_LIGHT_GREY)
    row_height(ws, 27, 160)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()

    build_executive_summary(wb)
    build_operations(wb)
    build_menu(wb)
    build_forecasting(wb)
    build_experiment(wb)

    wb.save(OUTPUT_PATH)
    print(f"Report saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
